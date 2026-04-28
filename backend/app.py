from __future__ import annotations

import json
import os
import sqlite3
import threading
import time
from datetime import datetime, timedelta
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

import requests
from dotenv import load_dotenv
from flask import Flask, jsonify, render_template, request, send_file, make_response
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / '.env')

DATA_DIR = Path(os.getenv('DATA_DIR', str(BASE_DIR / 'data'))).resolve()
DATA_DIR.mkdir(parents=True, exist_ok=True)

DB_PATH = DATA_DIR / 'submissions.db'
EXCEL_PATH = DATA_DIR / 'xodimlar.xlsx'
JSON_PATH = DATA_DIR / 'submissions.json'
TELEGRAM_OFFSET_PATH = DATA_DIR / 'telegram_offset.txt'

BOT_TOKEN = os.getenv('BOT_TOKEN', '').strip()
ADMIN_CHAT_IDS = [item.strip() for item in os.getenv('ADMIN_CHAT_IDS', '').split(',') if item.strip()]
SEND_EXCEL_TO_ADMIN_EACH_SUBMISSION = os.getenv('SEND_EXCEL_TO_ADMIN_EACH_SUBMISSION', 'false').strip().lower() == 'true'
TIMEZONE_OFFSET_HOURS = int(os.getenv('TIMEZONE_OFFSET_HOURS', '5'))
BOT_POLLING_ENABLED = os.getenv('BOT_POLLING_ENABLED', 'true').strip().lower() == 'true'
BOT_POLL_INTERVAL = float(os.getenv('BOT_POLL_INTERVAL', '2'))
FRONTEND_URL = os.getenv('FRONTEND_URL', '').strip()

_bot_polling_lock = threading.Lock()
_bot_polling_started = False

ADMIN_MENU_BUTTON = "📄 Excel formatda ko'rish"
TEACHER_POSITION = 'Ustoz'

REQUIRED_FIELDS = [
    'full_name', 'phone', 'branch', 'position', 'gender',
    'birth_date', 'work_start_year', 'work_start_month',
]

FIELD_NAMES = [
    'sent_at', 'full_name', 'phone', 'branch', 'position', 'teacher_subject', 'gender',
    'birth_date', 'work_start_year', 'work_start_month',
    'father_name', 'father_deceased', 'father_phone', 'father_birth_date',
    'mother_name', 'mother_deceased', 'mother_birth_date', 'mother_phone',
    'spouse_name', 'spouse_deceased', 'spouse_unmarried', 'spouse_divorced', 'spouse_birth_date', 'spouse_phone',
    'child1_name', 'child1_none', 'child1_birth_date',
    'child2_name', 'child2_none', 'child2_birth_date',
    'child3_name', 'child3_none', 'child3_birth_date',
    'father_in_law_name', 'father_in_law_deceased', 'father_in_law_birth_date',
    'mother_in_law_name', 'mother_in_law_deceased', 'mother_in_law_birth_date',
]

app = Flask(__name__)

# CORS
if FRONTEND_URL:
    CORS(
        app,
        resources={r"/api/*": {"origins": [FRONTEND_URL]}},
        allow_headers=["Content-Type", "Authorization"],
        methods=["GET", "POST", "OPTIONS"],
    )
else:
    CORS(
        app,
        resources={r"/api/*": {"origins": "*"}},
        allow_headers=["Content-Type", "Authorization"],
        methods=["GET", "POST", "OPTIONS"],
    )


@app.after_request
def add_cors_headers(response):
    origin = request.headers.get('Origin', '').strip()

    if FRONTEND_URL:
        if origin == FRONTEND_URL:
            response.headers['Access-Control-Allow-Origin'] = origin
    else:
        if origin:
            response.headers['Access-Control-Allow-Origin'] = origin
        else:
            response.headers['Access-Control-Allow-Origin'] = '*'

    response.headers['Vary'] = 'Origin'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    return response


def current_tashkent_time() -> str:
    return (datetime.utcnow().replace(microsecond=0) + timedelta(hours=TIMEZONE_OFFSET_HOURS)).strftime('%Y-%m-%d %H:%M:%S')


def parse_bool(payload: dict[str, Any], key: str) -> bool:
    return str(payload.get(key, 'false')).strip().lower() == 'true'


def pretty_bool(value: Any, positive: str = 'Ha', negative: str = "Yo'q") -> str:
    return positive if str(value).strip().lower() == 'true' else negative


def split_date_parts(value: Any) -> tuple[str, str, str]:
    raw = str(value or '').strip()
    if not raw:
        return '', '', ''
    for fmt in ('%Y-%m-%d', '%d.%m.%Y'):
        try:
            parsed = datetime.strptime(raw, fmt)
            return parsed.strftime('%Y'), parsed.strftime('%m'), parsed.strftime('%d')
        except ValueError:
            pass
    return '', '', ''


def get_db_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA journal_mode=WAL;')
    conn.execute('PRAGMA synchronous=NORMAL;')
    return conn


def initialize_database() -> None:
    with get_db_connection() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS submissions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sent_at TEXT NOT NULL,
                full_name TEXT,
                phone TEXT,
                branch TEXT,
                position TEXT,
                teacher_subject TEXT,
                gender TEXT,
                birth_date TEXT,
                work_start_year TEXT,
                work_start_month TEXT,
                father_name TEXT,
                father_deceased TEXT,
                father_phone TEXT,
                father_birth_date TEXT,
                mother_name TEXT,
                mother_deceased TEXT,
                mother_birth_date TEXT,
                mother_phone TEXT,
                spouse_name TEXT,
                spouse_deceased TEXT,
                spouse_unmarried TEXT,
                spouse_divorced TEXT,
                spouse_birth_date TEXT,
                spouse_phone TEXT,
                child1_name TEXT,
                child1_none TEXT,
                child1_birth_date TEXT,
                child2_name TEXT,
                child2_none TEXT,
                child2_birth_date TEXT,
                child3_name TEXT,
                child3_none TEXT,
                child3_birth_date TEXT,
                father_in_law_name TEXT,
                father_in_law_deceased TEXT,
                father_in_law_birth_date TEXT,
                mother_in_law_name TEXT,
                mother_in_law_deceased TEXT,
                mother_in_law_birth_date TEXT
            )
            """
        )
        conn.commit()


def validate_payload(payload: dict[str, Any]) -> str | None:
    for field in REQUIRED_FIELDS:
        if not str(payload.get(field, '')).strip():
            return f"Majburiy maydon to'ldirilmagan: {field}"
    if str(payload.get('position', '')).strip() == TEACHER_POSITION and not str(payload.get('teacher_subject', '')).strip():
        return "Ustoz uchun fan yo'nalishini tanlash majburiy."
    return None


def normalize_payload(payload: dict[str, Any]) -> dict[str, str]:
    is_unmarried = parse_bool(payload, 'spouse_unmarried')
    is_divorced = parse_bool(payload, 'spouse_divorced')
    return {
        'sent_at': current_tashkent_time(),
        'full_name': str(payload.get('full_name', '')).strip(),
        'phone': str(payload.get('phone', '')).strip(),
        'branch': str(payload.get('branch', '')).strip(),
        'position': str(payload.get('position', '')).strip(),
        'teacher_subject': str(payload.get('teacher_subject', '')).strip(),
        'gender': str(payload.get('gender', '')).strip(),
        'birth_date': str(payload.get('birth_date', '')).strip(),
        'work_start_year': str(payload.get('work_start_year', '')).strip(),
        'work_start_month': str(payload.get('work_start_month', '')).strip(),
        'father_name': str(payload.get('father_name', '')).strip(),
        'father_deceased': pretty_bool(payload.get('father_deceased'), 'Vafot etgan', 'Tirik'),
        'father_phone': str(payload.get('father_phone', '')).strip(),
        'father_birth_date': str(payload.get('father_birth_date', '')).strip(),
        'mother_name': str(payload.get('mother_name', '')).strip(),
        'mother_deceased': pretty_bool(payload.get('mother_deceased'), 'Vafot etgan', 'Tirik'),
        'mother_birth_date': str(payload.get('mother_birth_date', '')).strip(),
        'mother_phone': str(payload.get('mother_phone', '')).strip(),
        'spouse_name': str(payload.get('spouse_name', '')).strip(),
        'spouse_deceased': pretty_bool(payload.get('spouse_deceased'), 'Vafot etgan', "Yo'q"),
        'spouse_unmarried': pretty_bool(payload.get('spouse_unmarried'), 'Ha', "Yo'q"),
        'spouse_divorced': pretty_bool(payload.get('spouse_divorced'), 'Ha', "Yo'q"),
        'spouse_birth_date': str(payload.get('spouse_birth_date', '')).strip(),
        'spouse_phone': str(payload.get('spouse_phone', '')).strip(),
        'child1_name': str(payload.get('child1_name', '')).strip(),
        'child1_none': pretty_bool(payload.get('child1_none')),
        'child1_birth_date': str(payload.get('child1_birth_date', '')).strip(),
        'child2_name': str(payload.get('child2_name', '')).strip(),
        'child2_none': pretty_bool(payload.get('child2_none')),
        'child2_birth_date': str(payload.get('child2_birth_date', '')).strip(),
        'child3_name': str(payload.get('child3_name', '')).strip(),
        'child3_none': pretty_bool(payload.get('child3_none')),
        'child3_birth_date': str(payload.get('child3_birth_date', '')).strip(),
        'father_in_law_name': '' if (is_unmarried or is_divorced) else str(payload.get('father_in_law_name', '')).strip(),
        'father_in_law_deceased': '' if (is_unmarried or is_divorced) else pretty_bool(payload.get('father_in_law_deceased'), 'Vafot etgan', 'Tirik'),
        'father_in_law_birth_date': '' if (is_unmarried or is_divorced) else str(payload.get('father_in_law_birth_date', '')).strip(),
        'mother_in_law_name': '' if (is_unmarried or is_divorced) else str(payload.get('mother_in_law_name', '')).strip(),
        'mother_in_law_deceased': '' if (is_unmarried or is_divorced) else pretty_bool(payload.get('mother_in_law_deceased'), 'Vafot etgan', 'Tirik'),
        'mother_in_law_birth_date': '' if (is_unmarried or is_divorced) else str(payload.get('mother_in_law_birth_date', '')).strip(),
    }


def load_rows() -> list[dict[str, str]]:
    with get_db_connection() as conn:
        rows = conn.execute(f"SELECT {', '.join(FIELD_NAMES)} FROM submissions ORDER BY id DESC").fetchall()
    return [dict(row) for row in rows]


def append_row(row: dict[str, str]) -> int:
    with get_db_connection() as conn:
        conn.execute(
            f"INSERT INTO submissions ({', '.join(FIELD_NAMES)}) VALUES ({', '.join('?' for _ in FIELD_NAMES)})",
            [row[field] for field in FIELD_NAMES],
        )
        conn.commit()
        total = conn.execute('SELECT COUNT(*) FROM submissions').fetchone()[0]
    return int(total)


def build_excel(rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Xodimlar'
    headers = [
        '№', 'Yuborilgan vaqt', 'F.I.SH', 'Telefon', 'Filial', 'Lavozim', "Fan yo'nalishi", 'Jinsi',
        "Tug'ilgan yil", "Tug'ilgan oy", "Tug'ilgan kun", 'Ish boshlagan yil', 'Ish boshlagan oy',
        'Otasi F.I.SH', 'Otasi holati', 'Otasi telefoni', "Otasi tug'ilgan yil", "Otasi tug'ilgan oy", "Otasi tug'ilgan kun",
        'Onasi F.I.SH', 'Onasi holati', "Onasi tug'ilgan yil", "Onasi tug'ilgan oy", "Onasi tug'ilgan kun", 'Onasi telefoni',
        "Turmush o'rtog'i F.I.SH", "Turmush o'rtog'i vafot etgan", 'Turmush qurmaganman', 'Ajrashganman',
        "Turmush o'rtog'i tug'ilgan yil", "Turmush o'rtog'i tug'ilgan oy", "Turmush o'rtog'i tug'ilgan kun", "Turmush o'rtog'i telefoni",
        '1-farzand F.I.SH', "1-farzand yo'q", "1-farzand tug'ilgan yil", "1-farzand tug'ilgan oy", "1-farzand tug'ilgan kun",
        '2-farzand F.I.SH', "2-farzand yo'q", "2-farzand tug'ilgan yil", "2-farzand tug'ilgan oy", "2-farzand tug'ilgan kun",
        '3-farzand F.I.SH', "3-farzand yo'q", "3-farzand tug'ilgan yil", "3-farzand tug'ilgan oy", "3-farzand tug'ilgan kun",
        'Qaynota F.I.SH', 'Qaynota holati', "Qaynota tug'ilgan yil", "Qaynota tug'ilgan oy", "Qaynota tug'ilgan kun",
        'Qaynona F.I.SH', 'Qaynona holati', "Qaynona tug'ilgan yil", "Qaynona tug'ilgan oy", "Qaynona tug'ilgan kun",
    ]
    ws.append(headers)

    ordered_rows = list(reversed(rows))
    for index, item in enumerate(ordered_rows, start=1):
        employee_birth = split_date_parts(item.get('birth_date'))
        father_birth = split_date_parts(item.get('father_birth_date'))
        mother_birth = split_date_parts(item.get('mother_birth_date'))
        spouse_birth = split_date_parts(item.get('spouse_birth_date'))
        child1_birth = split_date_parts(item.get('child1_birth_date'))
        child2_birth = split_date_parts(item.get('child2_birth_date'))
        child3_birth = split_date_parts(item.get('child3_birth_date'))
        father_in_law_birth = split_date_parts(item.get('father_in_law_birth_date'))
        mother_in_law_birth = split_date_parts(item.get('mother_in_law_birth_date'))
        ws.append([
            index, item.get('sent_at', ''), item.get('full_name', ''), item.get('phone', ''), item.get('branch', ''),
            item.get('position', ''), item.get('teacher_subject', ''), item.get('gender', ''),
            employee_birth[0], employee_birth[1], employee_birth[2], item.get('work_start_year', ''), item.get('work_start_month', ''),
            item.get('father_name', ''), item.get('father_deceased', ''), item.get('father_phone', ''), father_birth[0], father_birth[1], father_birth[2],
            item.get('mother_name', ''), item.get('mother_deceased', ''), mother_birth[0], mother_birth[1], mother_birth[2], item.get('mother_phone', ''),
            item.get('spouse_name', ''), item.get('spouse_deceased', ''), item.get('spouse_unmarried', ''), item.get('spouse_divorced', ''),
            spouse_birth[0], spouse_birth[1], spouse_birth[2], item.get('spouse_phone', ''),
            item.get('child1_name', ''), item.get('child1_none', ''), child1_birth[0], child1_birth[1], child1_birth[2],
            item.get('child2_name', ''), item.get('child2_none', ''), child2_birth[0], child2_birth[1], child2_birth[2],
            item.get('child3_name', ''), item.get('child3_none', ''), child3_birth[0], child3_birth[1], child3_birth[2],
            item.get('father_in_law_name', ''), item.get('father_in_law_deceased', ''), father_in_law_birth[0], father_in_law_birth[1], father_in_law_birth[2],
            item.get('mother_in_law_name', ''), item.get('mother_in_law_deceased', ''), mother_in_law_birth[0], mother_in_law_birth[1], mother_in_law_birth[2],
        ])

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    for column in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 28)
    ws.freeze_panes = 'A2'

    rep = wb.create_sheet('Hisobot')
    rep['A1'] = 'Xodimlar hisobot'
    rep['A1'].font = Font(size=14, bold=True)
    rep['A3'] = 'Jami anketa'
    rep['A4'] = 'Ustozlar soni'
    rep['A5'] = 'Erkaklar soni'
    rep['A6'] = 'Ayollar soni'
    rep['A7'] = 'Oxirgi yuborilgan vaqt'
    for ref in ('A3', 'A4', 'A5', 'A6', 'A7'):
        rep[ref].font = Font(bold=True)
    rep['B3'] = len(rows)
    rep['B4'] = sum(1 for item in rows if item.get('position', '') == 'Ustoz')
    rep['B5'] = sum(1 for item in rows if item.get('gender', '') == 'Erkak')
    rep['B6'] = sum(1 for item in rows if item.get('gender', '') == 'Ayol')
    rep['B7'] = rows[0].get('sent_at', '') if rows else ''
    rep.column_dimensions['A'].width = 28
    rep.column_dimensions['B'].width = 22

    with NamedTemporaryFile(delete=False, suffix='.xlsx', dir=DATA_DIR) as tmp:
        temp_path = Path(tmp.name)
    try:
        wb.save(temp_path)
        temp_path.replace(EXCEL_PATH)
    finally:
        temp_path.unlink(missing_ok=True)


def sync_files() -> list[dict[str, str]]:
    rows = load_rows()
    JSON_PATH.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding='utf-8')
    build_excel(rows)
    return rows


def telegram_request(method: str, *, data: dict[str, Any] | None = None, files: dict[str, Any] | None = None) -> bool:
    if not BOT_TOKEN:
        return False
    try:
        response = requests.post(f'https://api.telegram.org/bot{BOT_TOKEN}/{method}', data=data, files=files, timeout=30)
        response.raise_for_status()
        payload = response.json()
        return bool(payload.get('ok'))
    except Exception:
        return False


def send_telegram_message(text: str) -> None:
    if not BOT_TOKEN or not ADMIN_CHAT_IDS:
        return
    for chat_id in ADMIN_CHAT_IDS:
        telegram_request('sendMessage', data={'chat_id': chat_id, 'text': text})


def send_excel_to_chat(chat_id: str, caption: str) -> bool:
    if not BOT_TOKEN or not EXCEL_PATH.exists():
        return False
    with EXCEL_PATH.open('rb') as fp:
        return telegram_request('sendDocument', data={'chat_id': chat_id, 'caption': caption}, files={'document': (EXCEL_PATH.name, fp)})


def send_excel_to_admins(caption: str) -> None:
    if not BOT_TOKEN or not ADMIN_CHAT_IDS or not EXCEL_PATH.exists():
        return
    for chat_id in ADMIN_CHAT_IDS:
        send_excel_to_chat(chat_id, caption)


def send_admin_menu(chat_id: str, text: str = "Kerakli bo'limni tanlang:") -> None:
    if not BOT_TOKEN:
        return
    keyboard = {
        'keyboard': [[{'text': ADMIN_MENU_BUTTON}]],
        'resize_keyboard': True,
        'one_time_keyboard': False,
    }
    telegram_request('sendMessage', data={
        'chat_id': chat_id,
        'text': text,
        'reply_markup': json.dumps(keyboard, ensure_ascii=False),
    })


def get_saved_offset() -> int:
    try:
        return int(TELEGRAM_OFFSET_PATH.read_text(encoding='utf-8').strip() or '0')
    except Exception:
        return 0


def save_offset(offset: int) -> None:
    TELEGRAM_OFFSET_PATH.write_text(str(offset), encoding='utf-8')


def handle_admin_message(message: dict[str, Any]) -> None:
    chat = message.get('chat') or {}
    chat_id = str(chat.get('id', '')).strip()
    if not chat_id or chat_id not in ADMIN_CHAT_IDS:
        return
    text = str(message.get('text', '')).strip()

    if text in ('/start', '/menu', '/help'):
        send_admin_menu(chat_id, "Assalomu alaykum. Excel faylni olish uchun pastdagi tugmani bosing.")
        return

    if text == ADMIN_MENU_BUTTON or text == '/excel':
        rows = sync_files()
        if rows:
            send_excel_to_chat(chat_id, f"Xodimlar ro'yxati. Jami: {len(rows)} ta")
        else:
            send_admin_menu(chat_id, "Hozircha saqlangan ma'lumot yo'q.")
        return

    send_admin_menu(chat_id, "Noma'lum buyruq. Pastdagi tugmani bosing.")


def bot_polling_loop() -> None:
    if not BOT_TOKEN or not ADMIN_CHAT_IDS or not BOT_POLLING_ENABLED:
        return
    offset = get_saved_offset()
    while True:
        try:
            response = requests.get(
                f'https://api.telegram.org/bot{BOT_TOKEN}/getUpdates',
                params={'timeout': 25, 'offset': offset, 'allowed_updates': ['message']},
                timeout=35,
            )
            response.raise_for_status()
            payload = response.json()
            if not payload.get('ok'):
                time.sleep(BOT_POLL_INTERVAL)
                continue
            for item in payload.get('result', []):
                offset = int(item['update_id']) + 1
                save_offset(offset)
                handle_admin_message(item.get('message') or {})
        except Exception:
            time.sleep(max(BOT_POLL_INTERVAL, 2))


def start_bot_polling() -> None:
    global _bot_polling_started
    if not BOT_TOKEN or not ADMIN_CHAT_IDS or not BOT_POLLING_ENABLED:
        return
    with _bot_polling_lock:
        if _bot_polling_started:
            return
        thread = threading.Thread(target=bot_polling_loop, daemon=True, name='telegram-bot-polling')
        thread.start()
        _bot_polling_started = True


def compute_stats(rows: list[dict[str, str]]) -> dict[str, int]:
    return {
        'total': len(rows),
        'teachers': sum(1 for row in rows if row.get('position') == 'Ustoz'),
        'male': sum(1 for row in rows if row.get('gender') == 'Erkak'),
        'female': sum(1 for row in rows if row.get('gender') == 'Ayol'),
    }


def startup() -> None:
    initialize_database()
    sync_files()
    start_bot_polling()


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/admin')
def admin_panel():
    rows = load_rows()
    return render_template('admin.html', rows=rows, stats=compute_stats(rows))


@app.route('/api/health', methods=['GET', 'OPTIONS'])
def health():
    if request.method == 'OPTIONS':
        return ('', 204)
    return jsonify({'success': True, 'message': 'OK'})


@app.route('/api/submit', methods=['POST', 'OPTIONS'])
def submit():
    if request.method == 'OPTIONS':
        return ('', 204)

    payload = request.get_json(silent=True)
    if not isinstance(payload, dict):
        return jsonify({'success': False, 'error': "Noto'g'ri JSON yuborildi."}), 400

    error = validate_payload(payload)
    if error:
        return jsonify({'success': False, 'error': error}), 400

    row = normalize_payload(payload)
    total_rows = append_row(row)
    rows = sync_files()

    message = (
        "Yangi xodim anketasi keldi\n"
        f"F.I.SH: {row['full_name']}\n"
        f"Telefon: {row['phone']}\n"
        f"Filial: {row['branch']}\n"
        f"Lavozim: {row['position']}\n"
        f"Fan: {row['teacher_subject'] or '-'}\n"
        f"Jinsi: {row['gender']}\n"
        f"Yuborilgan vaqt: {row['sent_at']}\n"
        f"Jami yozuvlar: {len(rows)}"
    )
    send_telegram_message(message)

    if SEND_EXCEL_TO_ADMIN_EACH_SUBMISSION:
        send_excel_to_admins(f"Yangilangan Excel fayl. Jami: {len(rows)} ta yozuv")

    return jsonify({'success': True, 'message': 'Saved', 'total_rows': total_rows})


@app.route('/api/submissions', methods=['GET', 'OPTIONS'])
def submissions():
    if request.method == 'OPTIONS':
        return ('', 204)
    return jsonify({'success': True, 'rows': load_rows()})


@app.route('/api/download/excel', methods=['GET', 'OPTIONS'])
def download_excel():
    if request.method == 'OPTIONS':
        return ('', 204)

    if not EXCEL_PATH.exists():
        sync_files()
    if not EXCEL_PATH.exists():
        return jsonify({'success': False, 'error': 'Excel topilmadi.'}), 404
    return send_file(EXCEL_PATH, as_attachment=True, download_name=EXCEL_PATH.name)


startup()

if __name__ == '__main__':
    port = int(os.getenv('PORT', '5000'))
    app.run(
        host='0.0.0.0',
        port=port,
        debug=os.getenv('FLASK_DEBUG', '0') == '1',
        use_reloader=os.getenv('FLASK_DEBUG', '0') == '1'
    )
